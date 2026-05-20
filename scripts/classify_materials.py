#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Deterministic material classifier based on product/class cards.

The script reads an Excel classifier with sheets:
- Class_Cards
- Terms
- Attributes
- Stop_Terms

Optional sheets are allowed:
- Values_Dictionary
- Positive_Examples
- Negative_Examples

It then classifies material names from CSV/XLSX using transparent Python scoring.
The process is repeatable: the same input and classifier always produce the same result.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


# -----------------------------------------------------------------------------
# Data models
# -----------------------------------------------------------------------------


@dataclass
class TermRule:
    class_code: str
    class_name: str
    term: str
    term_type: str = "keyword"
    match_type: str = "phrase"
    weight: float = 1.0
    is_required: bool = False
    comment: str = ""


@dataclass
class StopRule:
    class_code: str
    class_name: str
    stop_term: str
    match_type: str = "phrase"
    penalty: float = 1.0
    comment: str = ""


@dataclass
class AttributeRule:
    class_code: str
    class_name: str
    attribute_name: str
    aliases: List[str] = field(default_factory=list)
    unit: str = ""
    is_required: bool = False
    weight: float = 0.5
    value_pattern: str = ""
    normalization_rule: str = ""
    comment: str = ""


@dataclass
class ClassCard:
    class_code: str
    class_name: str
    class_path: str = ""
    main_product_name: str = ""
    search_text: str = ""


@dataclass
class CandidateScore:
    class_code: str
    class_name: str
    score: float
    raw_score: float
    matched_terms: List[str]
    matched_attributes: List[str]
    stop_terms: List[str]
    comment: str


# -----------------------------------------------------------------------------
# Normalization
# -----------------------------------------------------------------------------


def normalize_text(value: Any) -> str:
    """Normalize text for deterministic matching."""
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = text.replace("ё", "е")

    # Common Russian/Latin visual variants.
    replacements = {
        "×": "x",
        "х": "x",
        "*": "x",
        "–": "-",
        "—": "-",
        "−": "-",
        "\u00a0": " ",
        "м²": "мм2",  # only for matching tolerance; original values are not rewritten
        "мм²": "мм2",
        "кв.мм": "мм2",
        "кв мм": "мм2",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    # Decimal dot between digits is treated like comma for matching.
    text = re.sub(r"(?<=\d)\.(?=\d)", ",", text)

    # Normalize spaces.
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_bool(value: Any) -> bool:
    text = normalize_text(value)
    return text in {"yes", "y", "true", "1", "да", "истина", "обяз", "required", "mandatory"}


def parse_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def split_aliases(value: Any) -> List[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[;|,]", text)
    return [p.strip() for p in parts if p.strip()]


# -----------------------------------------------------------------------------
# Excel reading helpers
# -----------------------------------------------------------------------------


def get_header_map(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        key = normalize_text(cell.value)
        if key:
            headers[key] = idx
    return headers


def row_value(row, header_map: Dict[str, int], name: str, default: Any = "") -> Any:
    idx = header_map.get(normalize_text(name))
    if not idx:
        return default
    value = row[idx - 1].value
    return default if value is None else value


def sheet_rows(wb, sheet_name: str) -> Iterable[Tuple[List[Any], Dict[str, int]]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    header_map = get_header_map(ws)
    rows = []
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue
        rows.append((row, header_map))
    return rows


# -----------------------------------------------------------------------------
# Classifier loading
# -----------------------------------------------------------------------------


def load_class_cards(wb) -> Dict[str, ClassCard]:
    result: Dict[str, ClassCard] = {}
    for row, h in sheet_rows(wb, "Class_Cards"):
        class_code = str(row_value(row, h, "class_code", "")).strip()
        class_name = str(row_value(row, h, "class_name", "")).strip()
        if not class_code or not class_name:
            continue
        result[class_code] = ClassCard(
            class_code=class_code,
            class_name=class_name,
            class_path=str(row_value(row, h, "class_path", "")).strip(),
            main_product_name=str(row_value(row, h, "main_product_name", "")).strip(),
            search_text=str(row_value(row, h, "search_text", "")).strip(),
        )
    return result


def load_terms(wb) -> Dict[str, List[TermRule]]:
    result: Dict[str, List[TermRule]] = defaultdict(list)
    for row, h in sheet_rows(wb, "Terms"):
        class_code = str(row_value(row, h, "class_code", "")).strip()
        class_name = str(row_value(row, h, "class_name", "")).strip()
        term = str(row_value(row, h, "term", "")).strip()
        if not class_code or not term:
            continue
        result[class_code].append(
            TermRule(
                class_code=class_code,
                class_name=class_name,
                term=term,
                term_type=str(row_value(row, h, "term_type", "keyword")).strip() or "keyword",
                match_type=str(row_value(row, h, "match_type", "phrase")).strip() or "phrase",
                weight=parse_float(row_value(row, h, "weight", 1.0), 1.0),
                is_required=normalize_bool(row_value(row, h, "is_required", "")),
                comment=str(row_value(row, h, "comment", "")).strip(),
            )
        )
    return result


def load_stop_terms(wb) -> Dict[str, List[StopRule]]:
    result: Dict[str, List[StopRule]] = defaultdict(list)
    for row, h in sheet_rows(wb, "Stop_Terms"):
        class_code = str(row_value(row, h, "class_code", "")).strip()
        class_name = str(row_value(row, h, "class_name", "")).strip()
        stop_term = str(row_value(row, h, "stop_term", "")).strip()
        if not class_code or not stop_term:
            continue
        result[class_code].append(
            StopRule(
                class_code=class_code,
                class_name=class_name,
                stop_term=stop_term,
                match_type=str(row_value(row, h, "match_type", "phrase")).strip() or "phrase",
                penalty=parse_float(row_value(row, h, "penalty", 1.0), 1.0),
                comment=str(row_value(row, h, "comment", "")).strip(),
            )
        )
    return result


def load_attributes(wb) -> Dict[str, List[AttributeRule]]:
    result: Dict[str, List[AttributeRule]] = defaultdict(list)
    for row, h in sheet_rows(wb, "Attributes"):
        class_code = str(row_value(row, h, "class_code", "")).strip()
        class_name = str(row_value(row, h, "class_name", "")).strip()
        attribute_name = str(row_value(row, h, "attribute_name", "")).strip()
        if not class_code or not attribute_name:
            continue
        result[class_code].append(
            AttributeRule(
                class_code=class_code,
                class_name=class_name,
                attribute_name=attribute_name,
                aliases=split_aliases(row_value(row, h, "aliases", "")),
                unit=str(row_value(row, h, "unit", "")).strip(),
                is_required=normalize_bool(row_value(row, h, "is_required", "")),
                weight=parse_float(row_value(row, h, "weight", 0.5), 0.5),
                value_pattern=str(row_value(row, h, "value_pattern", "")).strip(),
                normalization_rule=str(row_value(row, h, "normalization_rule", "")).strip(),
                comment=str(row_value(row, h, "comment", "")).strip(),
            )
        )
    return result


# -----------------------------------------------------------------------------
# Matching and scoring
# -----------------------------------------------------------------------------


def term_matches(text_norm: str, term: str, match_type: str) -> bool:
    term_norm = normalize_text(term)
    if not term_norm:
        return False

    mt = normalize_text(match_type) or "phrase"

    if mt == "exact":
        return text_norm == term_norm

    if mt == "contains":
        return term_norm in text_norm

    if mt == "word":
        # Word boundary for Cyrillic, Latin, digits and symbols frequently used in markings.
        pattern = r"(?<![a-zа-я0-9])" + re.escape(term_norm) + r"(?![a-zа-я0-9])"
        return re.search(pattern, text_norm, flags=re.IGNORECASE) is not None

    if mt == "regex":
        try:
            return re.search(term, text_norm, flags=re.IGNORECASE) is not None
        except re.error:
            # Bad regex should not break the whole run; fallback to contains.
            return term_norm in text_norm

    # Default: phrase match with normalized whitespace.
    phrase = re.escape(term_norm).replace(r"\ ", r"\s+")
    pattern = r"(?<![a-zа-я0-9])" + phrase + r"(?![a-zа-я0-9])"
    return re.search(pattern, text_norm, flags=re.IGNORECASE) is not None


def attribute_matches(text_norm: str, attr: AttributeRule) -> bool:
    # 1) Direct regex pattern from classifier.
    if attr.value_pattern:
        try:
            if re.search(attr.value_pattern, text_norm, flags=re.IGNORECASE):
                return True
        except re.error:
            pass

    # 2) Alias/name presence together with unit or number.
    names = [attr.attribute_name] + attr.aliases
    names_norm = [normalize_text(x) for x in names if normalize_text(x)]
    unit_norm = normalize_text(attr.unit)

    has_name = any(n in text_norm for n in names_norm)
    if not has_name:
        return False

    if unit_norm and re.search(r"\d+[,.]?\d*\s*" + re.escape(unit_norm), text_norm):
        return True

    # If no unit is configured, alias presence itself can be useful but weaker.
    return bool(has_name and not unit_norm)


def score_class(
    material_text: str,
    card: ClassCard,
    term_rules: List[TermRule],
    stop_rules: List[StopRule],
    attr_rules: List[AttributeRule],
) -> CandidateScore:
    text_norm = normalize_text(material_text)

    raw = 0.0
    positive_capacity = 0.0
    matched_terms: List[str] = []
    matched_attributes: List[str] = []
    matched_stop_terms: List[str] = []

    # Class name / main product name fallback: useful when Terms are not complete yet.
    fallback_names = [card.main_product_name, card.class_name]
    for fallback in fallback_names:
        if fallback and term_matches(text_norm, fallback, "phrase"):
            raw += 0.35
            matched_terms.append(f"fallback:{fallback}")
            break

    required_terms = [t for t in term_rules if t.is_required]
    required_matched = 0

    for term in term_rules:
        weight = max(0.0, term.weight)
        positive_capacity += weight
        if term_matches(text_norm, term.term, term.match_type):
            raw += weight
            matched_terms.append(f"{term.term} [{term.term_type}:{weight:.2f}]")
            if term.is_required:
                required_matched += 1

    # Required terms: if they exist, at least one should normally match.
    if required_terms:
        if required_matched == 0:
            raw -= 0.50
        elif required_matched == len(required_terms):
            raw += 0.15

    # Attributes add supporting signal, capped to avoid overpowering keywords.
    attr_score = 0.0
    for attr in attr_rules:
        if attribute_matches(text_norm, attr):
            contribution = max(0.0, attr.weight) * 0.45
            attr_score += contribution
            matched_attributes.append(f"{attr.attribute_name}:{contribution:.2f}")

    raw += min(attr_score, 1.25)
    positive_capacity += min(sum(max(0.0, a.weight) * 0.45 for a in attr_rules), 1.25)

    # Stop terms are anti-signs.
    penalty_total = 0.0
    for stop in stop_rules:
        if term_matches(text_norm, stop.stop_term, stop.match_type):
            penalty = max(0.0, min(1.0, stop.penalty))
            penalty_total += penalty
            matched_stop_terms.append(f"{stop.stop_term}:{penalty:.2f}")

    raw -= penalty_total

    # Convert raw score to 0..1 in a stable, interpretable way.
    # Denominator is based on available positive signal, with a lower bound.
    denominator = max(1.0, min(positive_capacity, 5.0))
    score = max(0.0, min(1.0, raw / denominator))

    comment_parts = []
    if matched_terms:
        comment_parts.append("термины: " + "; ".join(matched_terms[:8]))
    if matched_attributes:
        comment_parts.append("атрибуты: " + "; ".join(matched_attributes[:6]))
    if matched_stop_terms:
        comment_parts.append("стоп-термины: " + "; ".join(matched_stop_terms[:6]))
    comment = " | ".join(comment_parts)

    return CandidateScore(
        class_code=card.class_code,
        class_name=card.class_name,
        score=round(score, 4),
        raw_score=round(raw, 4),
        matched_terms=matched_terms,
        matched_attributes=matched_attributes,
        stop_terms=matched_stop_terms,
        comment=comment,
    )


def classify_material(
    material_text: str,
    class_cards: Dict[str, ClassCard],
    terms_by_class: Dict[str, List[TermRule]],
    stops_by_class: Dict[str, List[StopRule]],
    attrs_by_class: Dict[str, List[AttributeRule]],
    top_k: int = 3,
) -> List[CandidateScore]:
    candidates: List[CandidateScore] = []

    for class_code, card in class_cards.items():
        candidate = score_class(
            material_text=material_text,
            card=card,
            term_rules=terms_by_class.get(class_code, []),
            stop_rules=stops_by_class.get(class_code, []),
            attr_rules=attrs_by_class.get(class_code, []),
        )
        if candidate.score > 0 or candidate.raw_score > 0 or candidate.stop_terms:
            candidates.append(candidate)

    candidates.sort(key=lambda c: (-c.score, -c.raw_score, c.class_code, c.class_name))
    return candidates[:top_k]


def decision_for(candidates: List[CandidateScore], accept_threshold: float, review_threshold: float, min_gap: float) -> Tuple[str, str]:
    if not candidates:
        return "no_class", "нет кандидатов с положительным скором"

    top = candidates[0]
    second_score = candidates[1].score if len(candidates) > 1 else 0.0
    gap = top.score - second_score

    if top.score >= accept_threshold and gap >= min_gap:
        return "auto_accept", f"top_score={top.score:.4f}; gap={gap:.4f}"

    if top.score >= review_threshold:
        return "review", f"нужна проверка: top_score={top.score:.4f}; gap={gap:.4f}"

    return "low_confidence", f"низкая уверенность: top_score={top.score:.4f}; gap={gap:.4f}"


# -----------------------------------------------------------------------------
# Input/output
# -----------------------------------------------------------------------------


def read_materials(path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_materials_xlsx(path, sheet_name)
    if suffix == ".csv":
        return read_materials_csv(path)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def find_column(headers: List[str], candidates: List[str]) -> Optional[str]:
    normalized = {normalize_text(h): h for h in headers}
    for candidate in candidates:
        key = normalize_text(candidate)
        if key in normalized:
            return normalized[key]
    return None


def read_materials_csv(path: Path) -> List[Dict[str, str]]:
    # Try semicolon first because Russian Excel exports often use it.
    text = path.read_text(encoding="utf-8-sig")
    sample = text[:4096]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    rows = []
    reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
    headers = reader.fieldnames or []
    id_col = find_column(headers, ["material_id", "id", "код", "номер", "n"])
    name_col = find_column(headers, ["material_name", "name", "clean_name", "наименование", "материал", "номенклатура"])
    if not name_col:
        raise ValueError("Cannot find material name column in CSV")

    for idx, row in enumerate(reader, start=1):
        rows.append({
            "material_id": str(row.get(id_col, idx) if id_col else idx),
            "material_name": str(row.get(name_col, "") or "").strip(),
        })
    return rows


def read_materials_xlsx(path: Path, sheet_name: Optional[str] = None) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    id_header = find_column(headers, ["material_id", "id", "код", "номер", "n"])
    name_header = find_column(headers, ["material_name", "name", "clean_name", "наименование", "материал", "номенклатура"])
    if not name_header:
        raise ValueError("Cannot find material name column in XLSX")

    id_idx = headers.index(id_header) + 1 if id_header else None
    name_idx = headers.index(name_header) + 1

    rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = row[name_idx - 1].value
        if name is None or not str(name).strip():
            continue
        material_id = row[id_idx - 1].value if id_idx else row_no - 1
        rows.append({
            "material_id": str(material_id),
            "material_name": str(name).strip(),
        })
    return rows


def build_output_rows(
    materials: List[Dict[str, str]],
    class_cards: Dict[str, ClassCard],
    terms_by_class: Dict[str, List[TermRule]],
    stops_by_class: Dict[str, List[StopRule]],
    attrs_by_class: Dict[str, List[AttributeRule]],
    top_k: int,
    accept_threshold: float,
    review_threshold: float,
    min_gap: float,
) -> List[Dict[str, Any]]:
    output_rows = []

    for material in materials:
        candidates = classify_material(
            material_text=material["material_name"],
            class_cards=class_cards,
            terms_by_class=terms_by_class,
            stops_by_class=stops_by_class,
            attrs_by_class=attrs_by_class,
            top_k=top_k,
        )
        decision, decision_comment = decision_for(candidates, accept_threshold, review_threshold, min_gap)

        row: Dict[str, Any] = {
            "material_id": material["material_id"],
            "material_name": material["material_name"],
            "decision": decision,
            "decision_comment": decision_comment,
        }

        for i in range(1, top_k + 1):
            if i <= len(candidates):
                c = candidates[i - 1]
                row[f"class_code_{i}"] = c.class_code
                row[f"class_name_{i}"] = c.class_name
                row[f"score_{i}"] = c.score
                row[f"raw_score_{i}"] = c.raw_score
                row[f"comment_{i}"] = c.comment
            else:
                row[f"class_code_{i}"] = ""
                row[f"class_name_{i}"] = ""
                row[f"score_{i}"] = ""
                row[f"raw_score_{i}"] = ""
                row[f"comment_{i}"] = ""

        output_rows.append(row)

    return output_rows


def write_output(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        write_csv(path, rows)
    elif suffix in {".xlsx", ".xlsm"}:
        write_xlsx(path, rows)
    else:
        raise ValueError(f"Unsupported output format: {path.suffix}")


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "classification_result"

    if not rows:
        wb.save(path)
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 80)

    wb.save(path)


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Deterministic material classification by product/class cards")
    parser.add_argument("--classifier", required=True, help="Path to classifier workbook XLSX")
    parser.add_argument("--input", required=True, help="Path to materials input CSV/XLSX")
    parser.add_argument("--output", required=True, help="Path to output CSV/XLSX")
    parser.add_argument("--sheet", default=None, help="Input XLSX sheet name. Default: first sheet")
    parser.add_argument("--top-k", type=int, default=3, help="Number of class candidates to output")
    parser.add_argument("--accept-threshold", type=float, default=0.55, help="Score threshold for auto_accept")
    parser.add_argument("--review-threshold", type=float, default=0.35, help="Score threshold for review")
    parser.add_argument("--min-gap", type=float, default=0.10, help="Minimum gap between candidate 1 and 2 for auto_accept")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    classifier_path = Path(args.classifier)
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not classifier_path.exists():
        print(f"ERROR: classifier file not found: {classifier_path}", file=sys.stderr)
        return 1
    if not input_path.exists():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        return 1

    wb = load_workbook(classifier_path, read_only=True, data_only=True)

    class_cards = load_class_cards(wb)
    terms_by_class = load_terms(wb)
    stops_by_class = load_stop_terms(wb)
    attrs_by_class = load_attributes(wb)

    if not class_cards:
        print("ERROR: no class cards loaded from sheet Class_Cards", file=sys.stderr)
        return 1

    materials = read_materials(input_path, args.sheet)
    rows = build_output_rows(
        materials=materials,
        class_cards=class_cards,
        terms_by_class=terms_by_class,
        stops_by_class=stops_by_class,
        attrs_by_class=attrs_by_class,
        top_k=max(1, args.top_k),
        accept_threshold=args.accept_threshold,
        review_threshold=args.review_threshold,
        min_gap=args.min_gap,
    )
    write_output(output_path, rows)

    print(json.dumps({
        "status": "ok",
        "materials": len(materials),
        "classes": len(class_cards),
        "output": str(output_path),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
